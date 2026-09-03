#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Anonimización del cuestionario aplicado a trabajadores — Proyecto SIMPA
Equipo AHMRV · ISR-401 · UTEQ

Lee la exportación cruda de Google Forms (.xlsx) y produce una copia
procesada en CSV sin las columnas identificativas, sin tocar el archivo
crudo. No inventa, imputa ni recalcula ninguna respuesta: solo elimina
columnas y descarta las columnas auxiliares "Puntos:" (sin uso analítico,
artefacto de que el formulario se aplicó con la plantilla "quiz" de Forms).

Principio: el archivo crudo (.xlsx) permanece intacto como evidencia
primaria; SOLO la salida de este script se declara como "dato procesado"
en el diccionario de datos (sección 6).

Uso:
    python3 anonimizar_encuesta.py <ruta_xlsx_crudo> [--out ruta_salida.csv]

Ejemplo real (desde la raíz del repo):
    python3 07_Publicacion/scripts/anonimizar_encuesta.py \\
        "02_Evidencias/Cuestionario/Respuestas/Sistema Inteligente de Mantenimiento de Palma Africana(1-62).xlsx" \\
        --out 07_Publicacion/respuestas_anonimizadas.csv

Requiere: openpyxl (pip install --break-system-packages openpyxl)
"""

import argparse
import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Falta la dependencia 'openpyxl'. Instale con: "
              "pip install --break-system-packages openpyxl")

# Columnas que identifican a la persona y se excluyen SIEMPRE del procesado.
# Se listan por nombre exacto de encabezado tal como los exporta Forms.
COLUMNAS_PII = {"Correo electrónico", "Nombre"}

# Prefijos de columnas auxiliares de la plantilla "quiz" de Forms, sin
# significado analítico para el estudio (no son PII; se excluyen por ruido).
PREFIJOS_SIN_USO_ANALITICO = ("Puntos: ", "Total de puntos",
                               "Hora de la última modificación")


def cargar_encabezados(ws):
    fila1 = next(ws.iter_rows(min_row=1, max_row=1))
    return [c.value for c in fila1]


def columnas_a_conservar(encabezados):
    indices = []
    excluidas_pii = []
    excluidas_ruido = []
    for i, nombre in enumerate(encabezados):
        nombre = nombre or ""
        if nombre in COLUMNAS_PII:
            excluidas_pii.append(nombre)
            continue
        if any(nombre.startswith(p) for p in PREFIJOS_SIN_USO_ANALITICO):
            excluidas_ruido.append(nombre)
            continue
        indices.append(i)
    return indices, excluidas_pii, excluidas_ruido


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx_crudo", type=Path)
    ap.add_argument("--out", type=Path,
                     default=Path("respuestas_anonimizadas.csv"))
    args = ap.parse_args()

    if not args.xlsx_crudo.exists():
        sys.exit(f"No existe el archivo crudo: {args.xlsx_crudo}")

    wb = openpyxl.load_workbook(args.xlsx_crudo, data_only=True)
    ws = wb.worksheets[0]

    encabezados = cargar_encabezados(ws)
    indices, excluidas_pii, excluidas_ruido = columnas_a_conservar(encabezados)

    filas_salida = [[encabezados[i] for i in indices]]
    n_filas = 0
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila is None or all(v is None for v in fila):
            continue
        filas_salida.append([fila[i] if i < len(fila) else None
                              for i in indices])
        n_filas += 1

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with open(args.out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerows(filas_salida)

    print(f"Archivo crudo:        {args.xlsx_crudo}")
    print(f"Filas de respuesta:   {n_filas}")
    print(f"Columnas conservadas: {len(indices)} de {len(encabezados)}")
    print(f"Columnas PII excluidas ({len(excluidas_pii)}): {excluidas_pii}")
    print(f"Columnas de ruido excluidas ({len(excluidas_ruido)}): "
          f"{excluidas_ruido}")
    print(f"Salida escrita en:    {args.out}")

    if not excluidas_pii:
        print("\n⚠️  ADVERTENCIA: no se encontró ninguna columna PII declarada "
              "('Correo electrónico', 'Nombre') en el encabezado. Verifique "
              "manualmente que el archivo crudo no cambió de estructura "
              "antes de asumir que ya no contiene datos identificables.")


if __name__ == "__main__":
    main()
